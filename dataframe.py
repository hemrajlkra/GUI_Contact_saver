from tkinter import *
from openpyxl import *
wb=load_workbook('C:\\Users\\hemra\\Downloads\\Documents\\details.xlsx')
# create the obj for sheet
sheet=wb.active
def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.cell(row=1, column=1).value="Name"
    sheet.cell(row=1, column=2).value="Username"
    sheet.cell(row=1, column=3).value="Email"
    sheet.cell(row=1, column=4).value = "Password"
#function to set focus

def focus1(event):
    name_field.focus_set()
def focus2(event):
    user_field.focus_set()
def focus3(event):
    email_field.focus_set()
def focus4(event):
    pass_field.focus_set()



def clear():
    name_field.delete(0,END)
    user_field.delete(0,END)
    email_field.delete(0,END)
    pass_field.delete(0,END)
def insert():
    if(name_field.get()== "" and
        user_field.get() == "" and
        email_field.get() == "" and
        pass_field.get() == ""
    ):
        print("Empty Input")
    else:
        current_row=sheet.max_row
        current_column=sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value=name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = user_field.get()
        sheet.cell(row=current_row + 1, column=3).value = email_field.get()
        sheet.cell(row=current_row + 1, column=4).value = pass_field.get()

        wb.save('C://Users//hemra//Downloads//Documents//details.xlsx')
        name_field.focus_set()
        clear()
if __name__ == '__main__':
    root=Tk()
    root.configure(background="light gray")
    root.title("User Details")
    root.geometry('500x300')
    excel()

    name_field = Entry(root)
    user_field = Entry(root)
    email_field = Entry(root)
    pass_field = Entry(root)

    name_field.grid(row=2, column=1, ipadx=100)
    user_field.grid(row=4, column=1, ipadx=100)
    email_field.grid(row=6, column=1, ipadx=100)
    pass_field.grid(row=8, column=1, ipadx=100)

    heading=Label(root,text="My Account Details", font = ('Comic Sans MS',15))
    heading.grid(row=0, column=1)

    Name=Label(root,text="Name",background="light green",width=10)
    Name.grid(row=2, column=0)
    Username = Label(root, text="Username", background="light green",width=10)
    Username.grid(row=4, column=0)
    email = Label(root, text="email", background="light green",width=10)
    email.grid(row=6, column=0)
    password = Label(root, text="Password", background="light green",width=10)
    password.grid(row=8, column=0)
    name_field.bind("<Return>",focus1)
    user_field.bind("<Return>",focus2)
    email_field.bind("<Return>", focus3)
    pass_field.bind("<Return>", focus4)
    excel()

    submit=Button(root,text="Save",fg='black',bg="red",command=insert,width=10, font = ('Comic Sans MS',10))
    submit.grid(row=12,column=1)
    root.mainloop()